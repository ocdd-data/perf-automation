"""
Regional Operational Data — fetches previous-month data and outputs a clean
Excel, then uploads it to Slack. Same style as kpi.py.

Sheets: SG, HK, NY, TH, KH, HCMC, HAN
Metrics per region/vehicle:
    Completed Trips, % promo trips, % non-promo trips, % match rate,
    GMV, Platform Fee Revenue, Average Fare

All query ids / params / columns below were verified against Redash and
compared to "Regional Operational Data.xlsx" (May-2026 matched ~exactly).

Run:
    python regional_operational_data.py            # previous month
    python regional_operational_data.py 2026-05    # a specific month (YYYY-MM)
"""

import os
import sys
from datetime import datetime, timedelta

sys.path.insert(0, os.path.abspath(os.path.join(os.path.dirname(__file__), "..")))

import pandas as pd
from dotenv import load_dotenv
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

from utils.helpers import Query, Redash
from utils.slack import SlackBot


# ---------------------------------------------------------------------------
# Date helpers
# ---------------------------------------------------------------------------

def month_info(month=None):
    """
    Return (date, start_date, end_date, label) for the report month.
    `month` is 'YYYY-MM'; if None, uses the previous calendar month.
    """
    if month:
        first = datetime.strptime(month + "-01", "%Y-%m-%d")
    else:
        today = datetime.today()
        first = (today.replace(day=1) - timedelta(days=1)).replace(day=1)

    # last day of that month
    nxt = (first.replace(day=28) + timedelta(days=4)).replace(day=1)
    last = nxt - timedelta(days=1)

    start_date = first.strftime("%Y-%m-%d")
    end_date = last.strftime("%Y-%m-%d")
    date = end_date                     # last day; matches the {date} convention
    label = first.strftime("%b_%Y")     # e.g. Jun_2026
    return date, start_date, end_date, label


def dr(start, end):
    return {"start": start, "end": end}


# ---------------------------------------------------------------------------
# Extraction helper — pull one value out of a Redash result DataFrame
# ---------------------------------------------------------------------------

def pull(df, column=None, num=None, den=None, filters=None,
         agg="first", scale=None, month=None):
    """
    filters: {col: value}  -- keep rows where col == value (case-insensitive).
             value "__MONTH__" matches rows whose date column == `month` (YYYY-MM).
    num/den: return num/den ratio (first row after filtering).
    agg:     "first" (row 0) or "sum" (sum the column across filtered rows).
    scale:   multiply the result (e.g. 0.01 to turn a percent into a fraction).
    """
    if df is None or not hasattr(df, "empty") or df.empty:
        return None

    d = df
    for col, val in (filters or {}).items():
        if col not in d.columns:
            return None
        if val == "__MONTH__":
            s = pd.to_datetime(d[col], errors="coerce").dt.strftime("%Y-%m")
            d = d[s == month]
        else:
            d = d[d[col].astype(str).str.upper() == str(val).upper()]
        if d.empty:
            return None

    try:
        if num and den:
            n = float(d.iloc[0][num])
            q = float(d.iloc[0][den])
            return (n / q) if q else None

        if column not in d.columns:
            return None
        if agg == "sum":
            v = pd.to_numeric(d[column], errors="coerce").sum()
        else:
            v = d.iloc[0][column]
        if pd.isna(v):
            return None
        v = float(v)
        return v * scale if scale else v
    except (TypeError, ValueError, KeyError):
        return None


def block(completed, promo, match, gmv, fee, avg_fare):
    """Ordered (metric_label, value) rows for one vehicle / region."""
    non_promo = (1 - promo) if promo is not None else None
    return [
        ("Completed Trips",      completed),
        ("% promo trips",        promo),
        ("% non-promo trips",    non_promo),
        ("% match rate",         match),
        ("GMV",                  gmv),
        ("Platform Fee Revenue", fee),
        ("Average Fare",         avg_fare),
    ]


# ---------------------------------------------------------------------------
# Excel output
# ---------------------------------------------------------------------------

HEADER_FILL  = PatternFill("solid", fgColor="1F3864")
VEHICLE_FILL = PatternFill("solid", fgColor="D9E1F2")
HEADER_FONT  = Font(name="Calibri", bold=True, color="FFFFFF", size=10)
VEHICLE_FONT = Font(name="Calibri", bold=True, size=10)
KPI_FONT     = Font(name="Calibri", size=10)


def write_sheet(ws, label, blocks):
    """
    blocks: list of (vehicle_label_or_None, [(metric_label, value), ...]).
    Single-vehicle regions pass vehicle_label = None.
    """
    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 22

    for c, txt in enumerate(["Metric", label], 1):
        cell = ws.cell(row=1, column=c, value=txt)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center" if c == 2 else "left")

    row = 2
    for vehicle, metrics in blocks:
        if vehicle:
            hc = ws.cell(row=row, column=1, value=vehicle)
            hc.font = VEHICLE_FONT
            hc.fill = VEHICLE_FILL
            ws.cell(row=row, column=2).fill = VEHICLE_FILL
            row += 1
        for label_, value in metrics:
            ws.cell(row=row, column=1, value=label_).font = KPI_FONT
            ws.cell(row=row, column=2, value=value).font = KPI_FONT
            row += 1
        row += 1   # blank line between vehicle blocks

    ws.freeze_panes = "B2"


# ---------------------------------------------------------------------------
# Region fetchers  (return list of (vehicle_or_None, block))
# ---------------------------------------------------------------------------

def fetch_sg(redash, date, s, e):
    redash.run_queries([
        Query(2183, params={"date": date}),
        Query(5382, params={"Date Range": dr(s, e)}),
        Query(6561, params={"date_range": dr(s, e), "region": "SG"}),
        Query(6189, params={"Date Range": dr(s, e), "region": "SG"}),
        Query(6708, params={"Date Range": dr(s, e), "region": "SG"}),
    ])
    q2183, q5382 = redash.get_result(2183), redash.get_result(5382)
    q6561, q6189, q6708 = redash.get_result(6561), redash.get_result(6189), redash.get_result(6708)
    return [(None, block(
        pull(q2183, column="completed"),
        pull(q5382, column="pct_of_promo_trips", filters={"region": "SG"}, scale=0.01),
        pull(q2183, num="matched", den="demand"),
        pull(q6561, column="total_gmv", agg="sum"),
        pull(q6189, column="total_system_fee", agg="sum"),
        pull(q6708, column="average_fare"),
    ))]


def fetch_hk(redash, date, s, e):
    redash.run_queries([
        Query(3771, params={"date_range": dr(s, e)}),
        Query(5382, params={"Date Range": dr(s, e)}),
        Query(6561, params={"date_range": dr(s, e), "region": "HK"}),
        Query(6189, params={"Date Range": dr(s, e), "region": "HK"}),
        Query(6708, params={"Date Range": dr(s, e), "region": "HK"}),
    ])
    q3771, q5382 = redash.get_result(3771), redash.get_result(5382)
    q6561, q6189, q6708 = redash.get_result(6561), redash.get_result(6189), redash.get_result(6708)
    return [(None, block(
        pull(q3771, column="completed"),
        pull(q5382, column="pct_of_promo_trips", filters={"region": "HK"}, scale=0.01),
        pull(q3771, num="matched", den="demand"),
        pull(q6561, column="total_gmv", agg="sum"),
        pull(q6189, column="total_system_fee", agg="sum"),
        pull(q6708, column="average_fare"),
    ))]


def fetch_ny(redash, date, s, e):
    # NY data lags ~1 day; run on/after the 2nd of the month for a full month.
    # NOTE: 7644 (GMV) filters on the month-START date, so it takes `s`, not `date`.
    redash.run_queries([
        Query(7579, params={"date": date}),            # nyc_completed
        Query(7578, params={"date": date}),            # matched / demand
        Query(7670, params={"Date Range": dr(s, e)}),  # promo %
        Query(7655, params={"Date Range": dr(s, e)}),  # system fee
        Query(7665, params={"Date Range": dr(s, e)}),  # average fare
        Query(7644, params={"date": s}),               # GMV (month-start date)
    ])
    q7579, q7578 = redash.get_result(7579), redash.get_result(7578)
    q7670, q7655, q7665 = redash.get_result(7670), redash.get_result(7655), redash.get_result(7665)
    q7644 = redash.get_result(7644)
    return [(None, block(
        pull(q7579, column="nyc_completed"),
        pull(q7670, column="pct_of_promo_trips", filters={"region": "NY"}, scale=0.01),
        pull(q7578, num="matched", den="demand"),
        pull(q7644, column="gmv"),                      # 7644 NY GMV
        pull(q7655, column="total_system_fee", agg="sum"),
        pull(q7665, column="average_fare"),
    ))]


def fetch_th(redash, date, s, e):
    redash.run_queries([
        Query(6577, params={"Date Range": dr(s, e), "region": "TH"}),
        Query(3106, params={"date": date}),
        Query(6565, params={"Date Range": dr(s, e), "Region": "TH"}),
        Query(6561, params={"date_range": dr(s, e), "region": "TH"}),
        Query(6189, params={"Date Range": dr(s, e), "region": "TH"}),
        Query(6564, params={"date": dr(s, e)}),
    ])
    q6577, q3106 = redash.get_result(6577), redash.get_result(3106)
    q6565, q6561 = redash.get_result(6565), redash.get_result(6561)
    q6189, q6564 = redash.get_result(6189), redash.get_result(6564)

    blocks = []
    for veh, wheel, grp, fare_col, num, den in [
        ("BIKE (2W)", "2W", "BIKE", "bike_average_fare", "bike_matched", "bike_demand"),
        ("CAR (4W)",  "4W", "CAR",  "car_average_fare",  "_4w_matched",  "_4w_demand"),
    ]:
        blocks.append((veh, block(
            pull(q6577, column="completed_rides", filters={"vehicle_type": wheel}),
            pull(q6565, column="pct_of_promo_trips", filters={"car_group": grp}, scale=0.01),
            pull(q3106, num=num, den=den),
            pull(q6561, column="total_gmv", filters={"car_group": grp}),
            pull(q6189, column="total_system_fee", filters={"vehicle_type": wheel}, agg="sum"),
            pull(q6564, column=fare_col),
        )))
    return blocks


def fetch_kh(redash, date, s, e):
    redash.run_queries([
        Query(6577, params={"Date Range": dr(s, e), "region": "KH"}),
        Query(6640, params={"Date Range": dr(s, e), "region": "KH"}),
        Query(6565, params={"Date Range": dr(s, e), "Region": "KH"}),
        Query(6561, params={"date_range": dr(s, e), "region": "KH"}),
        Query(6189, params={"Date Range": dr(s, e), "region": "KH"}),
        Query(6708, params={"Date Range": dr(s, e), "region": "KH"}),
    ])
    q6577, q6640 = redash.get_result(6577), redash.get_result(6640)
    q6565, q6561 = redash.get_result(6565), redash.get_result(6561)
    q6189, q6708 = redash.get_result(6189), redash.get_result(6708)

    blocks = []
    for veh, wheel, grp in [
        ("TUKTUK (3W)", "3W", "TUKTUK"),
        ("CAR (4W)",    "4W", "CAR"),
        ("BIKE (2W)",   "2W", "BIKE"),
    ]:
        blocks.append((veh, block(
            pull(q6577, column="completed_rides", filters={"vehicle_type": wheel}),
            pull(q6565, column="pct_of_promo_trips", filters={"car_group": grp}, scale=0.01),
            pull(q6640, column="match_rate", filters={"vehicle_type": wheel}),
            pull(q6561, column="total_gmv", filters={"car_group": grp}),
            pull(q6189, column="total_system_fee", filters={"vehicle_type": wheel}, agg="sum"),
            pull(q6708, column="average_fare", filters={"vehicle_type": wheel}),
        )))
    return blocks


def fetch_vn_city(redash, date, s, e, month, city):
    """Vietnam city (HCM / HAN). region param = VN; rows filtered by city."""
    redash.run_queries([
        Query(7666, params={"date_range": dr(s, e)}),
        Query(6562, params={"Date Range": dr(s, e)}),
        Query(6640, params={"Date Range": dr(s, e), "region": "VN"}),
        Query(6189, params={"Date Range": dr(s, e), "region": "VN"}),
        Query(6708, params={"Date Range": dr(s, e), "region": "VN"}),
    ])
    q7666, q6562 = redash.get_result(7666), redash.get_result(6562)
    q6640, q6189, q6708 = redash.get_result(6640), redash.get_result(6189), redash.get_result(6708)

    blocks = []
    for veh, wheel, grp in [("BIKE (2W)", "2W", "BIKE"), ("CAR (4W)", "4W", "CAR")]:
        gmv_f = {"ride_month": "__MONTH__", "city": city, "car_group": grp}
        blocks.append((veh, block(
            pull(q7666, column="completed_rides", filters=gmv_f, month=month),
            pull(q6562, column="pct_of_promo_trips", filters={"city": city, "car_group": grp}, scale=0.01),
            pull(q6640, column="match_rate", filters={"city": city, "vehicle_type": wheel}),
            pull(q7666, column="total_gmv", filters=gmv_f, month=month),
            pull(q6189, column="total_system_fee", filters={"city": city, "vehicle_type": wheel}, agg="sum"),
            pull(q6708, column="average_fare", filters={"city": city, "vehicle_type": wheel}),
        )))
    return blocks


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    load_dotenv()

    redash = Redash(
        key=os.getenv("REDASH_API_KEY"),
        base_url=os.getenv("REDASH_BASE_URL"),
    )

    month_arg = sys.argv[1] if len(sys.argv) > 1 else None     # optional YYYY-MM
    date, start_date, end_date, label = month_info(month_arg)
    month = start_date[:7]

    print(f"Regional Operational Data - {label}")
    print(f"  Period : {start_date} -> {end_date}\n")

    tasks = [
        ("SG",   lambda: fetch_sg(redash, date, start_date, end_date)),
        ("HK",   lambda: fetch_hk(redash, date, start_date, end_date)),
        ("NY",   lambda: fetch_ny(redash, date, start_date, end_date)),
        ("TH",   lambda: fetch_th(redash, date, start_date, end_date)),
        ("KH",   lambda: fetch_kh(redash, date, start_date, end_date)),
        ("HCMC", lambda: fetch_vn_city(redash, date, start_date, end_date, month, "HCM")),
        ("HAN",  lambda: fetch_vn_city(redash, date, start_date, end_date, month, "HAN")),
    ]

    wb = Workbook()
    wb.remove(wb.active)

    for name, fn in tasks:
        print(f"-> {name}")
        try:
            blocks = fn()
            ws = wb.create_sheet(name)
            write_sheet(ws, label, blocks)
            print("   done")
        except Exception as e:
            import traceback
            print(f"   FAILED: {e}")
            traceback.print_exc()
            ws = wb.create_sheet(name)
            ws["A1"] = f"FAILED: {e}"

    output_file = f"Regional_Operational_Data_{label}.xlsx"
    wb.save(output_file)
    print(f"\nSaved: {output_file}")

    slack = SlackBot()
    slack.uploadFile(
        output_file,
        os.getenv("SLACK_CHANNEL"),
        f"Regional Operational Data for {label.replace('_', ' ')}",
    )
    print("Uploaded to Slack")

    try:
        os.remove(output_file)
    except OSError:
        pass


if __name__ == "__main__":
    main()
