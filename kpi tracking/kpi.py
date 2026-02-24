"""
KPI Tracker — fetches previous month data and outputs a clean Excel.
Sheets: SG, HK, TH, VN-HCMC, VN-Hanoi, KH-PNH, KH-OTHERS
"""

import os
from datetime import datetime, timedelta
import sys

sys.path.insert(0, os.path.abspath(os.path.join(os.path.dirname(__file__), "..")))

import pandas as pd
from dotenv import load_dotenv
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

from utils.constants import IDS, TIMEZONES
from utils.helpers import Query, Redash
from utils.slack import SlackBot


# ---------------------------------------------------------------------------
# Date helpers
# ---------------------------------------------------------------------------

def prev_month_info():
    today       = datetime.today()
    last_day    = today.replace(day=1) - timedelta(days=1)
    date        = last_day.strftime("%Y-%m-%d")
    start_date  = last_day.replace(day=1).strftime("%Y-%m-%d")
    end_date    = date
    label       = last_day.strftime("%b_%Y")
    two_back    = last_day.replace(day=1) - timedelta(days=1)
    churn_start = two_back.replace(day=1).strftime("%Y-%m-%d")
    return date, start_date, end_date, churn_start, label


def dr(start, end):
    return {"start": start, "end": end}


def pick_month_row(df, month_start):
    """
    Pick the row matching the report month.
    month_start is 'YYYY-MM-01' string.
    """
    if df is None or not hasattr(df, "empty") or df.empty:
        return pd.DataFrame()

    target = pd.to_datetime(month_start).date()

    for col in ["month", "month_start", "date"]:
        if col in df.columns:
            s = pd.to_datetime(df[col], errors="coerce").dt.date
            exact = df[s == target]
            if not exact.empty:
                return exact.head(1)

            # fallback: same YYYY-MM
            s2 = pd.to_datetime(df[col], errors="coerce").astype(str)
            mask = s2.str.startswith(month_start[:7])
            fallback = df[mask]
            if not fallback.empty:
                return fallback.head(1)

    return df.head(1)


# ---------------------------------------------------------------------------
# Data helpers
# ---------------------------------------------------------------------------

def safe_val(df, col, default=None):
    try:
        if df is None or not hasattr(df, "empty") or df.empty or col not in df.columns:
            return default
        val = df[col].iloc[0]
        return None if pd.isna(val) else val
    except Exception:
        return default


def filter_rows(df, **kwargs):
    """Filter df by col==val (case-insensitive). Returns first matching row."""
    if df is None or not hasattr(df, "empty") or df.empty:
        return pd.DataFrame()
    mask = pd.Series([True] * len(df), index=df.index)
    for col, val in kwargs.items():
        if col in df.columns:
            mask &= df[col].astype(str).str.upper() == str(val).upper()
    out = df[mask]
    return out.head(1) if not out.empty else pd.DataFrame()

def vt(df, vehicle_type, city=None):
    """
    Filter df by vehicle type across different schemas:
      - vehicle_type
      - w_type
      - wheel_group
      - vehicle
    Also filters by city if provided AND df has a city column.
    """
    if df is None or not hasattr(df, "empty") or df.empty:
        return pd.DataFrame()

    # find the correct type column
    type_col = None
    for c in ["vehicle_type", "w_type", "wheel_group", "vehicle"]:
        if c in df.columns:
            type_col = c
            break

    kw = {}
    if type_col:
        kw[type_col] = vehicle_type

    if city and "city" in df.columns:
        kw["city"] = city

    return filter_rows(df, **kw)


def has_col(df, col):
    return df is not None and hasattr(df, "columns") and col in df.columns


def div(num, denom):
    try:
        if num is None or denom is None or float(denom) == 0:
            return None
        return float(num) / float(denom)
    except Exception:
        return None

def first_present_val(df, candidates, default=None):
    """Return first non-null value among candidate columns from a 1-row df."""
    if df is None or not hasattr(df, "empty") or df.empty:
        return default
    for c in candidates:
        if c in df.columns:
            v = df[c].iloc[0]
            if not pd.isna(v):
                return v
    return default


def filter_city_if_possible(df, city):
    """If df has city column, filter by city; else return df as-is."""
    if df is None or not hasattr(df, "empty") or df.empty:
        return pd.DataFrame()
    if "city" in df.columns:
        return filter_rows(df, city=city)
    return df

def filter_all(df, **kwargs):
    if df is None or not hasattr(df, "empty") or df.empty:
        return pd.DataFrame()
    mask = pd.Series(True, index=df.index)
    for col, val in kwargs.items():
        if col in df.columns:
            mask &= df[col].astype(str).str.upper() == str(val).upper()
    return df[mask]

# ---------------------------------------------------------------------------
# Excel output
# ---------------------------------------------------------------------------

HEADER_FILL  = PatternFill("solid", fgColor="1F3864")
SECTION_FILL = PatternFill("solid", fgColor="D9E1F2")
HEADER_FONT  = Font(name="Calibri", bold=True, color="FFFFFF", size=10)
SECTION_FONT = Font(name="Calibri", bold=True, size=10)
KPI_FONT     = Font(name="Calibri", size=10)
VAL_FONT     = Font(name="Calibri", size=10)


def write_sheet(ws, month_label, sections):
    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 45
    ws.column_dimensions["D"].width = 22

    for c, txt in enumerate(["Type", "Section", "KPI", month_label], 1):
        cell = ws.cell(row=1, column=c, value=txt)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")

    row = 2
    for vehicle_type, section_name, kpi_label, value in sections:
        if section_name is not None and row > 2:
            row += 1
        c_vehicle = ws.cell(row=row, column=1, value=vehicle_type or "")
        c_section = ws.cell(row=row, column=2, value=section_name or "")
        c_kpi     = ws.cell(row=row, column=3, value=kpi_label)
        c_val     = ws.cell(row=row, column=4, value=value)
        if section_name:
            for c in (c_vehicle, c_section, c_kpi, c_val):
                c.fill = SECTION_FILL
                c.font = SECTION_FONT
        else:
            c_kpi.font = KPI_FONT
            c_val.font = VAL_FONT
        row += 1

    ws.freeze_panes = "D2"


# ---------------------------------------------------------------------------
# Row templates
# ---------------------------------------------------------------------------

def rows_sg_hk(d):
    """SG/HK — 4W only, full Unit Economics including Median Searched Fare."""
    return [
        ("4W", "Demand",          "Unique Rider Searches (Daily Average)",          d.get("unique_rider_searches")),
        (None,  None,             "Unique Bookings Created (Daily Average)",        d.get("unique_bookings_daily")),
        (None,  None,             "Completed Trips (Daily Average)",                d.get("completed_daily")),
        (None,  None,             "Book-Search Ratio",                              d.get("book_search_ratio")),
        (None,  None,             "Completion Rate",                                d.get("completion_rate")),
        (None,  "Supply",         "Daily Avg Drivers with >= 1 Ping",              d.get("pinged_drivers_daily")),
        (None,  None,             "Average Driver Online Hours",                    d.get("avg_online_hours")),
        (None,  None,             "Monthly Unique Completed Drivers",               d.get("completed_drivers")),
        (None,  None,             "Avg Completed Trip Per Driver",                  d.get("ride_per_driver")),
        (None,  None,             "Driver Utilisation (utilised / online)",         d.get("driver_utilisation")),
        (None,  "Efficiency",     "Match Rate",                                     d.get("match_rate")),
        (None,  None,             "First Try Cater Rate",                           d.get("first_try_cater_rate")),
        (None,  None,             "Median Time to Match (Seconds)",                 d.get("median_time_to_match_sec")),
        (None,  None,             "Median Pick Up ETA (Minutes)",                   d.get("median_eta")),
        (None,  "Unit Economics", "Median Searched Fare",                           d.get("median_searched_fare")),
        (None,  None,             "Median Booked Fare",                             d.get("median_booked_fare")),
        (None,  None,             "Median Completed Fare (Rider)",                  d.get("median_completed_fare")),
        (None,  None,             "Median Matched Trip Driver Earnings",             d.get("median_matched_driver_earnings")),
        (None,  None,             "Median Completed Driver Earnings",               d.get("median_completed_driver_earnings")),
        (None,  "Spend",          "Driver Incentive Cost / Trip",                   d.get("driver_incentive")),
        (None,  None,             "Rider Promo Cost / Trip",                        d.get("rider_promo_cost")),
        (None,  None,             "Rider Promo Rate",                               d.get("rider_promo_rate")),
        (None,  None,             "Non-trip Marketing Cost",                        d.get("non_trip_marketing")),
        (None,  None,             "Total Spend (Driver+Rider) / Trip",              d.get("total_spend")),
        (None,  None,             "Rider+Driver System Fee (Monthly Average)",      d.get("system_fee")),
        (None,  "Retention",      "Rider Completed",                                d.get("rider_completed")),
        (None,  None,             "Rider Activated",                                d.get("rider_activated")),
        (None,  None,             "Rider Churn Rate",                               d.get("rider_churn_rate")),
        (None,  None,             "Driver Completed",                               d.get("driver_completed")),
        (None,  None,             "Driver Activated",                               d.get("driver_activated")),
        (None,  None,             "Driver Churn Rate",                              d.get("driver_churn_rate")),
    ]


def vehicle_block(vtype, d):
    """
    Vehicle block for TH/VN/KH.
    Median Searched Fare is in General only — NOT here.
    """
    return [
        (vtype, "Demand",          "Unique Bookings Created (Daily Average)",        d.get("unique_bookings_daily")),
        (None,  None,             "Completed Trips (Daily Average)",                d.get("completed_daily")),
        (None,  None,             "Completion Rate",                                d.get("completion_rate")),
        (None,  "Supply",         "Daily Avg Drivers with >= 1 Ping",              d.get("pinged_drivers_daily")),
        (None,  None,             "Average Driver Online Hours",                    d.get("avg_online_hours")),
        (None,  None,             "Monthly Completed Drivers",                      d.get("completed_drivers")),
        (None,  None,             "Avg Completed Trip Per Driver",                  d.get("ride_per_driver")),
        (None,  None,             "Driver Utilisation (utilised / online)",         d.get("driver_utilisation")),
        (None,  "Efficiency",     "Match Rate",                                     d.get("match_rate")),
        (None,  None,             "First Try Cater Rate",                           d.get("first_try_cater_rate")),
        (None,  None,             "Daily Average Median Time to Match (seconds)",   d.get("median_time_to_match_sec")),
        (None,  None,             "Median Pick Up ETA",                             d.get("median_eta")),
        (None,  "Unit Economics", "Median Booked Fare",                             d.get("median_booked_fare")),
        (None,  None,             "Median Completed Fare",                          d.get("median_completed_fare")),
        (None,  None,             "Median Matched Trip Driver Earnings",             d.get("median_matched_driver_earnings")),
        (None,  None,             "Median Completed Driver Earnings",               d.get("median_completed_driver_earnings")),
        (None,  "Spend",          "Driver Incentive Cost / Trip",                   d.get("driver_incentive")),
        (None,  None,             "Rider Promo Cost / Trip",                        d.get("rider_promo_cost")),
        (None,  None,             "Rider Promo Rate",                               d.get("rider_promo_rate")),
        (None,  None,             "Total Spend / Trip",                             d.get("total_spend")),
        (None,  None,             "Rider+Driver System Fee (Monthly Average)",      d.get("system_fee")),
        (None,  "Retention",      "Rider Completed",                                d.get("rider_completed")),
        (None,  None,             "Rider Activated",                                d.get("rider_activated")),
        (None,  None,             "Rider Churn Rate",                               d.get("rider_churn_rate")),
        (None,  None,             "Driver Completed",                               d.get("driver_completed")),
        (None,  None,             "Driver Activated",                               d.get("driver_activated")),
        (None,  None,             "Driver Churn Rate",                              d.get("driver_churn_rate")),
    ]


def rows_th_vn(gen, d2w, d4w):
    return [
        ("General", None, "Unique Rider Searches",   gen.get("unique_rider_searches")),
        (None, None,      "Book-Search Ratio",        gen.get("book_search_ratio")),
        (None, None,      "Median Searched Fare",     gen.get("median_searched_fare")),
        (None, None,      "Non-trip Marketing Cost",  gen.get("non_trip_marketing")),
    ] + vehicle_block("2W", d2w) + vehicle_block("4W", d4w)


def rows_kh(gen, d2w, d3w, d4w, has_2w=True):
    out = [
        ("General", None, "Unique Rider Searches",   gen.get("unique_rider_searches")),
        (None, None,      "Book-Search Ratio",        gen.get("book_search_ratio")),
        (None, None,      "Median Searched Fare",     gen.get("median_searched_fare")),
        (None, None,      "Non-trip Marketing Cost",  gen.get("non_trip_marketing")),
    ]
    if has_2w:
        out += vehicle_block("2W", d2w)
    out += vehicle_block("3W", d3w)
    out += vehicle_block("4W", d4w)
    return out


# ---------------------------------------------------------------------------
# SG — confirmed working
# ---------------------------------------------------------------------------

def fetch_sg(redash, date, start_date, end_date, churn_start):
    region    = "SG"
    region_id = IDS["SG"]

    queries = [
        Query(2183, params={"date": date}),
        Query(2189, params={"date": date}),   # rider activated
        Query(2194, params={"date": date}),
        Query(2198, params={"date": date}),
        Query(2204, params={"date": date}),
        Query(2206, params={"date": date}),   # driver activated
        Query(2208, params={"date": date}),
        Query(2209, params={"date": date}),
        Query(2210, params={"date": date}),
        Query(4691, params={"date": date}),
        Query(4814, params={"date_range": dr(start_date, end_date), "region": region_id}),
        Query(4819, params={"date_range": dr(start_date, end_date), "region": region}),
        Query(6139, params={"Date Range": dr(start_date, end_date), "region": region}),  # driver incentive
        Query(6152, params={"Date Range": dr(start_date, end_date), "region": region}),
        Query(6138, params={"Date Range": dr(start_date, end_date), "region": region, "city": "SIN"}),
        Query(6030, params={"Date Range": dr(start_date, end_date), "region": region}),
        Query(6189, params={"Date Range": dr(start_date, end_date), "region": region}),
    ]
    redash.run_queries(queries)

    df23  = redash.get_result(4819)
    df1   = redash.get_result(2183)
    df5   = redash.get_result(2189)
    df6   = redash.get_result(2194)
    bq4   = redash.get_result(2198)
    df14  = redash.get_result(2204)
    df15  = redash.get_result(2206)
    bq5   = redash.get_result(2208)
    df16  = redash.get_result(2209)
    df8   = redash.get_result(2210)
    df19  = redash.get_result(4691)
    df22  = redash.get_result(4814)
    df2   = redash.get_result(6139)
    q6152 = redash.get_result(6152)
    q6138 = redash.get_result(6138)
    q6030 = redash.get_result(6030)
    q6189 = redash.get_result(6189)

    days = int(date.split("-")[2])

    return {
        "unique_rider_searches":            safe_val(df23, "rider_unique_search_daily_avg"),
        "unique_bookings_daily":            safe_val(q6152, "unique_orders_daily_avg"),
        "completed_daily":                  div(safe_val(df1, "completed"), days),
        "book_search_ratio":                safe_val(df23, "book_search_ratio_daily"),
        "completion_rate":                  div(safe_val(df1, "completed"), safe_val(df1, "demand")),
        "pinged_drivers_daily":             safe_val(bq4, "pinged_drivers_daily"),
        "avg_online_hours":                 safe_val(bq5, "avg_online_hour"),
        "completed_drivers":                safe_val(df1, "completed_drivers"),
        "ride_per_driver":                  safe_val(df14, "ride_per_driver"),
        "driver_utilisation":               div(safe_val(df16, "avg_utilisation_hours"), safe_val(bq5, "avg_online_hour")),
        "match_rate":                       div(safe_val(df1, "matched"), safe_val(df1, "demand")),
        "first_try_cater_rate":             safe_val(df19, "first_try_cater_rate"),
        "median_time_to_match_sec":         safe_val(df22, "median_time_to_match_sec"),
        "median_eta":                       safe_val(df8, "median_eta"),
        "median_searched_fare":             safe_val(q6138, "median_searched_fare"),
        "median_booked_fare":               safe_val(q6138, "median_booked_fare"),
        "median_completed_fare":            safe_val(q6138, "median_completed_fare"),
        "median_matched_driver_earnings":   safe_val(q6138, "median_matched_driver_earnings"),
        "median_completed_driver_earnings": safe_val(q6138, "median_completed_driver_earnings"),
        "driver_incentive":                 safe_val(df2, "incentive_per_completed_trip"),
        "rider_promo_cost":                 safe_val(q6030, "rider_promo_cost_per_trip"),
        "rider_promo_rate":                 safe_val(q6030, "rider_promo_rate"),
        "non_trip_marketing":               None,
        "total_spend":                      None,
        "system_fee":                       safe_val(q6189, "avg_system_fee_per_trip"),
        "rider_completed":                  safe_val(df6, "completed_monthly"),
        "rider_activated":                  safe_val(df5, "all_time"),
        "rider_churn_rate":                 None,
        "driver_completed":                 safe_val(df1, "completed_drivers"),
        "driver_activated":                 safe_val(df15, "all_time"),
        "driver_churn_rate":                None,
    }


# ---------------------------------------------------------------------------
# HK — confirmed working
# ---------------------------------------------------------------------------

def fetch_hk(redash, date, start_date, end_date, churn_start):
    region    = "HK"
    region_id = IDS["HK"]

    queries = [
        Query(3771, params={"date_range": dr(start_date, end_date)}),
        Query(3774, params={"date_range": dr(start_date, end_date)}),
        Query(3779, params={"date_range": dr(start_date, end_date)}),
        Query(3780, params={"date_range": dr(start_date, end_date)}),
        Query(3781, params={"date_range": dr(start_date, end_date)}),   # driver activated
        Query(3782, params={"date_range": dr(start_date, end_date)}),
        Query(3783, params={"date_range": dr(start_date, end_date)}),
        Query(3787, params={"date_range": dr(start_date, end_date)}),   # rider activated
        Query(4753, params={"date_range": dr(start_date, end_date)}),
        Query(4814, params={"date_range": dr(start_date, end_date), "region": region_id}),
        Query(4819, params={"date_range": dr(start_date, end_date), "region": region}),
        Query(6139, params={"Date Range": dr(start_date, end_date), "region": region}),  # driver incentive
        Query(6152, params={"Date Range": dr(start_date, end_date), "region": region}),
        Query(6138, params={"Date Range": dr(start_date, end_date), "region": region, "city": "HKG"}),
        Query(6030, params={"Date Range": dr(start_date, end_date), "region": region}),
        Query(6189, params={"Date Range": dr(start_date, end_date), "region": region}),
    ]
    redash.run_queries(queries)

    df1   = redash.get_result(3771)
    df3   = redash.get_result(3774)
    bq3   = redash.get_result(3779)
    df5   = redash.get_result(3780)
    df6   = redash.get_result(3781)
    bq4   = redash.get_result(3782)
    df7   = redash.get_result(3783)
    df11  = redash.get_result(3787)
    df14  = redash.get_result(4753)
    df15  = redash.get_result(4814)
    df16  = redash.get_result(4819)
    df2   = redash.get_result(6139)
    q6152 = redash.get_result(6152)
    q6138 = redash.get_result(6138)
    q6030 = redash.get_result(6030)
    q6189 = redash.get_result(6189)

    days = int(date.split("-")[2])

    return {
        "unique_rider_searches":            safe_val(df16, "rider_unique_search_daily_avg"),
        "unique_bookings_daily":            safe_val(q6152, "unique_orders_daily_avg"),
        "completed_daily":                  div(safe_val(df1, "completed"), days),
        "book_search_ratio":                safe_val(df16, "book_search_ratio_daily"),
        "completion_rate":                  div(safe_val(df1, "completed_phv"), safe_val(df1, "demand_phv")),
        "pinged_drivers_daily":             safe_val(bq3, "pinged_drivers_daily"),
        "avg_online_hours":                 safe_val(bq4, "avg_online_hour"),
        "completed_drivers":                safe_val(df1, "completed_drivers"),
        "ride_per_driver":                  safe_val(df5, "ride_per_driver"),
        "driver_utilisation":               div(safe_val(df7, "avg_utilisation_hours"), safe_val(bq4, "avg_online_hour")),
        "match_rate":                       div(safe_val(df1, "matched"), safe_val(df1, "demand")),
        "first_try_cater_rate":             safe_val(df14, "first_try_cater_rate"),
        "median_time_to_match_sec":         safe_val(df15, "median_time_to_match_sec"),
        "median_eta":                       safe_val(df3, "median_eta"),
        "median_searched_fare":             safe_val(q6138, "median_searched_fare"),
        "median_booked_fare":               safe_val(q6138, "median_booked_fare"),
        "median_completed_fare":            safe_val(q6138, "median_completed_fare"),
        "median_matched_driver_earnings":   safe_val(q6138, "median_matched_driver_earnings"),
        "median_completed_driver_earnings": safe_val(q6138, "median_completed_driver_earnings"),
        "driver_incentive":                 safe_val(df2, "incentive_per_completed_trip"),
        "rider_promo_cost":                 safe_val(q6030, "rider_promo_cost_per_trip"),
        "rider_promo_rate":                 safe_val(q6030, "rider_promo_rate"),
        "non_trip_marketing":               None,
        "total_spend":                      None,
        "system_fee":                       safe_val(q6189, "avg_system_fee_per_trip"),
        "rider_completed":                  safe_val(df1, "completed_riders"),
        "rider_activated":                  safe_val(df11, "all_time"),
        "rider_churn_rate":                 None,
        "driver_completed":                 safe_val(df1, "completed_drivers"),
        "driver_activated":                 safe_val(df6, "all_time"),
        "driver_churn_rate":                None,
    }


# ---------------------------------------------------------------------------
# TH
# ---------------------------------------------------------------------------

def fetch_th(redash, date, start_date, end_date, churn_start):
    region    = "TH"
    region_id = IDS["TH"]
    timezone  = TIMEZONES["TH"]

    queries = [
        Query(3106, params={"date": date}),
        Query(3117, params={"region": region, "timezone": timezone, "date": date}),   # driver activated
        Query(3122, params={"region": region,    "timezone": timezone, "date": date}),
        Query(3123, params={"region": region,    "timezone": timezone, "date": date}),
        Query(3125, params={"region": region_id, "timezone": timezone, "date": date}),
        Query(3126, params={"region": region_id, "timezone": timezone, "date": date}),
        Query(3127, params={"region": region,    "timezone": timezone, "date": date}),
        Query(3128, params={"region": region,    "timezone": timezone, "date": date}),
        Query(3130, params={"region": region_id, "timezone": timezone, "date": date}),
        Query(3131, params={"region": region_id, "timezone": timezone, "date": date}),
        Query(4814, params={"date_range": dr(start_date, end_date), "region": region_id}),
        Query(4819, params={"date_range": dr(start_date, end_date), "region": region}),
        Query(6145, params={"Date Range": dr(start_date, end_date), "region": region, "city": "BKK"}),
        Query(6152, params={"Date Range": dr(start_date, end_date), "region": region}),
        Query(6144, params={"Date Range": dr(start_date, end_date), "region": region}),
        Query(4509, params={"Date Range": dr(start_date, end_date), "region": region}),
        Query(6148, params={"Date Range": dr(start_date, end_date), "region": region}),
        Query(6138, params={"Date Range": dr(start_date, end_date), "region": region, "city": "BKK"}),
        Query(6030, params={"Date Range": dr(start_date, end_date), "region": region}),
        Query(6189, params={"Date Range": dr(start_date, end_date), "region": region}),
        Query(6366, params={"Date Range": dr(start_date, end_date), "region": region, "city": "BKK"}),  # median searched fare
        Query(6143, params={"Date Range": dr(churn_start, end_date)}),                # churn
    ]
    redash.run_queries(queries)

    df1    = redash.get_result(3106)
    df_act = redash.get_result(3117)
    df17   = redash.get_result(3122)
    df18   = redash.get_result(3123)
    bq7    = redash.get_result(3125)
    bq8    = redash.get_result(3126)
    df19   = redash.get_result(3127)
    df20   = redash.get_result(3128)
    bq10   = redash.get_result(3130)
    bq11   = redash.get_result(3131)
    q6145  = redash.get_result(6145)
    q6152  = redash.get_result(6152)
    q6144  = redash.get_result(6144)
    q4509  = redash.get_result(4509)
    q6148  = redash.get_result(6148)
    q6138  = redash.get_result(6138)
    q6030  = redash.get_result(6030)
    q6189  = redash.get_result(6189)
    q6366  = redash.get_result(6366)
    q6143  = redash.get_result(6143)

    # split by vehicle_type
    q6152_2w = vt(q6152, "2W");  q6152_4w = vt(q6152, "4W")
    q6144_2w = vt(q6144, "2W");  q6144_4w = vt(q6144, "4W")
    q4509_2w = vt(q4509, "2W");  q4509_4w = vt(q4509, "4W")
    q6148_2w = vt(q6148, "2W");  q6148_4w = vt(q6148, "4W")
    q6138_2w = vt(q6138, "2W");  q6138_4w = vt(q6138, "4W")
    q6189_2w = vt(q6189, "2W");  q6189_4w = vt(q6189, "4W")
    q6143_2w = vt(q6143, "2W");  q6143_4w = vt(q6143, "4W")
    q6030_2w = vt(q6030, "2W", city="BKK")
    q6030_4w = vt(q6030, "4W", city="BKK")

    general = {
        "unique_rider_searches": safe_val(q6145, "avg_daily_searches"),
        "book_search_ratio":     safe_val(q6145, "monthly_book_search_ratio"),
        "median_searched_fare":  safe_val(q6366, "median_searched_fare"),
        "non_trip_marketing":    None,
    }

    d2w = {
        "unique_bookings_daily":            safe_val(q6152_2w, "unique_orders_daily_avg"),
        "completed_daily":                  safe_val(q6152_2w, "completed_daily_avg"),
        "completion_rate":                  safe_val(q6152_2w, "completion_rate"),
        "pinged_drivers_daily":             safe_val(bq7,  "pinged_bike_daily"),
        "avg_online_hours":                 safe_val(bq8,  "avg_online_hour"),
        "completed_drivers":                safe_val(df1,  "bike_completed_drivers"),
        "ride_per_driver":                  safe_val(df17, "ride_per_bike"),
        "driver_utilisation":               div(safe_val(df18, "avg_utilisation_hours"), safe_val(bq8, "avg_online_hour")),
        "match_rate":                       div(safe_val(df1, "bike_matched"), safe_val(df1, "bike_demand")),
        "first_try_cater_rate":             safe_val(q6144_2w, "first_try_cater_rate"),
        "median_time_to_match_sec":         safe_val(q4509_2w, "median_time_to_match_sec"),
        "median_eta":                       safe_val(q6148_2w, "median_eta_minutes"),
        "median_booked_fare":               safe_val(q6138_2w, "median_booked_fare"),
        "median_completed_fare":            safe_val(q6138_2w, "median_completed_fare"),
        "median_matched_driver_earnings":   safe_val(q6138_2w, "median_matched_driver_earnings"),
        "median_completed_driver_earnings": safe_val(q6138_2w, "median_completed_driver_earnings"),
        "driver_incentive":                 None,
        "rider_promo_cost":                 safe_val(q6030_2w, "rider_promo_cost_per_trip"),
        "rider_promo_rate":                 safe_val(q6030_2w, "rider_promo_rate"),
        "total_spend":                      None,
        "system_fee":                       safe_val(q6189_2w, "avg_system_fee_per_trip"),
        "rider_completed":                  safe_val(df1, "bike_completed_riders"),
        "rider_activated":                  safe_val(q6143_2w, "rider_activated_2w"),
        "rider_churn_rate":                 safe_val(q6143_2w, "rider_churn_rate_2w"),
        "driver_completed":                 safe_val(df1, "bike_completed_drivers"),
        "driver_activated":                 safe_val(df_act, "bike_all_time"),
        "driver_churn_rate":                None,   # from perf sheet
    }

    d4w = {
        "unique_bookings_daily":            safe_val(q6152_4w, "unique_orders_daily_avg"),
        "completed_daily":                  safe_val(q6152_4w, "completed_daily_avg"),
        "completion_rate":                  safe_val(q6152_4w, "completion_rate"),
        "pinged_drivers_daily":             safe_val(bq10, "pinged_4w_daily"),
        "avg_online_hours":                 safe_val(bq11, "avg_online_hour"),
        "completed_drivers":                safe_val(df1,  "_4w_completed_drivers"),
        "ride_per_driver":                  safe_val(df19, "ride_per_4w"),
        "driver_utilisation":               div(safe_val(df20, "avg_utilisation_hours"), safe_val(bq11, "avg_online_hour")),
        "match_rate":                       div(safe_val(df1, "_4w_matched"), safe_val(df1, "_4w_demand")),
        "first_try_cater_rate":             safe_val(q6144_4w, "first_try_cater_rate"),
        "median_time_to_match_sec":         safe_val(q4509_4w, "median_time_to_match_sec"),
        "median_eta":                       safe_val(q6148_4w, "median_eta_minutes"),
        "median_booked_fare":               safe_val(q6138_4w, "median_booked_fare"),
        "median_completed_fare":            safe_val(q6138_4w, "median_completed_fare"),
        "median_matched_driver_earnings":   safe_val(q6138_4w, "median_matched_driver_earnings"),
        "median_completed_driver_earnings": safe_val(q6138_4w, "median_completed_driver_earnings"),
        "driver_incentive":                 None,
        "rider_promo_cost":                 safe_val(q6030_4w, "rider_promo_cost_per_trip"),
        "rider_promo_rate":                 safe_val(q6030_4w, "rider_promo_rate"),
        "total_spend":                      None,
        "system_fee":                       safe_val(q6189_4w, "avg_system_fee_per_trip"),
        "rider_completed":                  safe_val(df1, "_4w_completed_riders"),
        "rider_activated":                  safe_val(q6143_4w, "rider_activated_4w"),
        "rider_churn_rate":                 safe_val(q6143_4w, "rider_churn_rate_4w"),
        "driver_completed":                 safe_val(df1, "_4w_completed_drivers"),
        "driver_activated":                 safe_val(df_act, "_4w_all_time"),
        "driver_churn_rate":                None,   # from perf sheet
    }

    return general, d2w, d4w


# ---------------------------------------------------------------------------
# VN (HCM / HAN)
# ---------------------------------------------------------------------------

def fetch_vn(redash, start_date, end_date, churn_start, city_code):
    region = "VN"

    queries = [
        # old-style (city scoped)
        Query(4562, params={"date_range": dr(start_date, end_date), "city": city_code}),  # rides
        Query(4566, params={"date_range": dr(start_date, end_date), "city": city_code}),  # ETA (kept, but not used for median_eta)
        Query(4570, params={"date_range": dr(start_date, end_date), "city": city_code}),  # drivers ping
        Query(4571, params={"date_range": dr(start_date, end_date), "city": city_code}),  # drivers daily
        Query(4572, params={"date_range": dr(start_date, end_date), "city": city_code}),  # drivers FT
        Query(4574, params={"date_range": dr(start_date, end_date), "city": city_code}),  # avg online hours
        Query(4576, params={"date_range": dr(start_date, end_date), "city": city_code}),  # utilisation
        Query(4580, params={"date_range": dr(start_date, end_date), "city": city_code}),  # rider FT
        Query(4581, params={"date_range": dr(start_date, end_date), "city": city_code}),  # rider daily

        # new-style
        Query(6145, params={"Date Range": dr(start_date, end_date), "region": region, "city": city_code}),  # searches + BSR (city param)
        Query(6152, params={"Date Range": dr(start_date, end_date), "region": region}),                     # region-only -> filter by city+vtype
        Query(6144, params={"Date Range": dr(start_date, end_date), "region": region}),                     # region-only -> filter by city+vtype
        Query(4509, params={"Date Range": dr(start_date, end_date), "region": region}),                     # region-only -> filter by city+vtype
        Query(6148, params={"Date Range": dr(start_date, end_date), "region": region}),                     # region-only -> filter by city+vtype
        Query(6138, params={"Date Range": dr(start_date, end_date), "region": region, "city": city_code}),  # fares (city param)
        Query(6030, params={"Date Range": dr(start_date, end_date), "region": region}),                     # promo region-only -> filter by city(+vtype)
        Query(6189, params={"Date Range": dr(start_date, end_date), "region": region}),                     # fee region-only -> filter by city(+vtype)
        Query(6366, params={"Date Range": dr(start_date, end_date), "region": region, "city": city_code}),  # searched fare (city param)

        # churn (same logic as TH)
        Query(6143, params={"Date Range": dr(churn_start, end_date)}),                                      # churn table
    ]
    redash.run_queries(queries)

    # old results
    df1    = redash.get_result(4562)
    bq3    = redash.get_result(4570)
    df5    = redash.get_result(4571)
    df6    = redash.get_result(4572)
    bq4    = redash.get_result(4574)
    df7    = redash.get_result(4576)
    df11   = redash.get_result(4580)
    df12   = redash.get_result(4581)

    # new results
    q6145  = redash.get_result(6145)
    q6152  = redash.get_result(6152)
    q6144  = redash.get_result(6144)
    q4509  = redash.get_result(4509)
    q6148  = redash.get_result(6148)
    q6138  = redash.get_result(6138)
    q6030  = redash.get_result(6030)
    q6189  = redash.get_result(6189)
    q6366  = redash.get_result(6366)
    q6143  = redash.get_result(6143)

    # IMPORTANT: region-only tables -> filter by city + vehicle_type
    q6152_2w = vt(q6152, "2W", city=city_code)
    q6152_4w = vt(q6152, "4W", city=city_code)

    q6148_2w = vt(q6148, "2W", city=city_code)
    q6148_4w = vt(q6148, "4W", city=city_code)

    q6144_2w = vt(q6144, "2W", city=city_code)
    q6144_4w = vt(q6144, "4W", city=city_code)

    q4509_2w = vt(q4509, "2W", city=city_code)
    q4509_4w = vt(q4509, "4W", city=city_code)

    # promo (region-only) -> city + vtype
    q6030_2w = vt(q6030, "2W", city=city_code)
    q6030_4w = vt(q6030, "4W", city=city_code)

    # system fee (region-only) -> city (+ vtype if present)
    if has_col(q6189, "vehicle_type") or has_col(q6189, "w_type"):
        q6189_2w = vt(q6189, "2W", city=city_code)
        q6189_4w = vt(q6189, "4W", city=city_code)
    else:
        q6189_2w = filter_rows(q6189, city=city_code)
        q6189_4w = filter_rows(q6189, city=city_code)

    # churn table -> pick target month row for the report month
    churn_city = pick_month_row(filter_rows(q6143, city=city_code), start_date)

    general = {
        "unique_rider_searches": safe_val(q6145, "avg_daily_searches"),
        "book_search_ratio":     safe_val(q6145, "monthly_book_search_ratio"),
        "median_searched_fare":  safe_val(q6366, "median_searched_fare"),
        "non_trip_marketing":    None,
    }

    d2w = {
        "unique_bookings_daily":            safe_val(q6152_2w, "unique_orders_daily_avg"),
        "completed_daily":                  safe_val(q6152_2w, "completed_daily_avg"),
        "completion_rate":                  safe_val(q6152_2w, "completion_rate"),

        "pinged_drivers_daily":             safe_val(bq3, "pinged_drivers_daily_bike"),
        "avg_online_hours":                 safe_val(bq4, "avg_online_hour_bike"),

        "completed_drivers":                safe_val(df1, "completed_drivers_bike"),
        "ride_per_driver":                  safe_val(df5, "ride_per_driver_bike"),
        "driver_utilisation":               div(safe_val(df7, "avg_utilisation_hours_bike"), safe_val(bq4, "avg_online_hour_bike")),

        "match_rate":                       div(safe_val(df1, "matched_bike"), safe_val(df1, "demand_bike")),
        "first_try_cater_rate":             safe_val(q6144_2w, "first_try_cater_rate"),
        "median_time_to_match_sec":         safe_val(q4509_2w, "median_time_to_match_sec"),

        # FIX: use 6148 consistently
        "median_eta":                       safe_val(q6148_2w, "median_eta_minutes"),

        "median_booked_fare":               safe_val(vt(q6138, "2W"), "median_booked_fare"),
        "median_completed_fare":            safe_val(vt(q6138, "2W"), "median_completed_fare"),
        "median_matched_driver_earnings":   safe_val(vt(q6138, "2W"), "median_matched_driver_earnings"),
        "median_completed_driver_earnings": safe_val(vt(q6138, "2W"), "median_completed_driver_earnings"),

        "driver_incentive":                 None,
        "rider_promo_cost":                 safe_val(q6030_2w, "rider_promo_cost_per_trip"),
        "rider_promo_rate":                 safe_val(q6030_2w, "rider_promo_rate"),
        "total_spend":                      None,
        "system_fee":                       safe_val(q6189_2w, "avg_system_fee_per_trip"),

        "rider_completed":                  safe_val(df12, "completed_monthly_bike"),
        "rider_activated":                  safe_val(df11, "all_time_bike"),
        "rider_churn_rate":                 None,

        "driver_completed":                 safe_val(df1, "completed_drivers_bike"),
        "driver_activated":                 safe_val(df6, "all_time_bike"),
        "driver_churn_rate":                None,
    }

    d4w = {
        "unique_bookings_daily":            safe_val(q6152_4w, "unique_orders_daily_avg"),
        "completed_daily":                  safe_val(q6152_4w, "completed_daily_avg"),
        "completion_rate":                  safe_val(q6152_4w, "completion_rate"),

        "pinged_drivers_daily":             safe_val(bq3, "pinged_drivers_daily_phv"),
        "avg_online_hours":                 safe_val(bq4, "avg_online_hour_phv"),

        "completed_drivers":                safe_val(df1, "completed_drivers_phv"),
        "ride_per_driver":                  safe_val(df5, "ride_per_driver_phv"),
        "driver_utilisation":               div(safe_val(df7, "avg_utilisation_hours_phv"), safe_val(bq4, "avg_online_hour_phv")),

        "match_rate":                       div(safe_val(df1, "matched_phv"), safe_val(df1, "demand_phv")),
        "first_try_cater_rate":             safe_val(q6144_4w, "first_try_cater_rate"),
        "median_time_to_match_sec":         safe_val(q4509_4w, "median_time_to_match_sec"),

        # FIX: use 6148 consistently
        "median_eta":                       safe_val(q6148_4w, "median_eta_minutes"),

        "median_booked_fare":               safe_val(vt(q6138, "4W"), "median_booked_fare"),
        "median_completed_fare":            safe_val(vt(q6138, "4W"), "median_completed_fare"),
        "median_matched_driver_earnings":   safe_val(vt(q6138, "4W"), "median_matched_driver_earnings"),
        "median_completed_driver_earnings": safe_val(vt(q6138, "4W"), "median_completed_driver_earnings"),

        "driver_incentive":                 None,
        "rider_promo_cost":                 safe_val(q6030_4w, "rider_promo_cost_per_trip"),
        "rider_promo_rate":                 safe_val(q6030_4w, "rider_promo_rate"),
        "total_spend":                      None,
        "system_fee":                       safe_val(q6189_4w, "avg_system_fee_per_trip"),

        "rider_completed":                  safe_val(df1, "completed_riders_phv"),
        "rider_activated":                  safe_val(df11, "all_time_phv"),
        "rider_churn_rate":                 None,

        "driver_completed":                 safe_val(df1, "completed_drivers_phv"),
        "driver_activated":                 safe_val(df6, "all_time_phv"),
        "driver_churn_rate":                None,
    }

    return general, d2w, d4w

# ---------------------------------------------------------------------------
# KH (PNH and KH-OTHERS)
# ---------------------------------------------------------------------------

def fetch_kh_city(redash, start_date, end_date, churn_start, city):
    region = "KH"

    queries = [
        Query(6145, params={"Date Range": dr(start_date, end_date), "region": region, "city": city}),
        Query(6366, params={"Date Range": dr(start_date, end_date), "region": region, "city": city}),

        Query(6152, params={"Date Range": dr(start_date, end_date), "region": region}),  # region-only -> filter city+vtype
        Query(6151, params={"Date Range": dr(start_date, end_date)}),
        Query(6157, params={"Date Range": dr(start_date, end_date)}),
        Query(6149, params={"Date Range": dr(start_date, end_date), "region": region}),
        Query(6161, params={"Date Range": dr(start_date, end_date)}),

        Query(6147, params={"Date Range": dr(start_date, end_date), "region": region}),
        Query(6144, params={"Date Range": dr(start_date, end_date), "region": region}),
        Query(4509, params={"Date Range": dr(start_date, end_date), "region": region}),
        Query(6148, params={"Date Range": dr(start_date, end_date), "region": region}),

        Query(6138, params={"Date Range": dr(start_date, end_date), "region": region, "city": city}),
        Query(6030, params={"Date Range": dr(start_date, end_date), "region": region}),  # region-only -> filter city(+vtype)
        Query(6189, params={"Date Range": dr(start_date, end_date), "region": region}),  # region-only -> filter city(+vtype)

        # CAR tables run across churn range, pick report month row
        Query(4659, params={"Date Range": dr(churn_start, end_date)}),
        Query(6150, params={"Date Range": dr(churn_start, end_date)}),
    ]
    redash.run_queries(queries)

    q6145 = redash.get_result(6145)
    q6366 = redash.get_result(6366)
    q6152 = redash.get_result(6152)
    q6151 = redash.get_result(6151)
    q6157 = redash.get_result(6157)
    q6149 = redash.get_result(6149)
    q6161 = redash.get_result(6161)
    q6147 = redash.get_result(6147)
    q6144 = redash.get_result(6144)
    q4509 = redash.get_result(4509)
    q6148 = redash.get_result(6148)
    q6138 = redash.get_result(6138)
    q6030 = redash.get_result(6030)
    q6189 = redash.get_result(6189)
    q4659 = redash.get_result(4659)
    q6150 = redash.get_result(6150)

    # ----- has_2w MUST be city-scoped -----
    # q6152_city = filter_city_if_possible(q6152, city)
    # avail = set()
    # if q6152_city is not None and not q6152_city.empty and has_col(q6152_city, "vehicle_type"):
    #     avail = set(q6152_city["vehicle_type"].astype(str).str.upper().unique())
    # has_2w = "2W" in avail 

    q6152_city = filter_rows(q6152, city=city) if has_col(q6152, "city") else q6152
    avail = set(q6152_city["vehicle_type"].astype(str).str.upper().unique()) if has_col(q6152_city, "vehicle_type") else set()
    has_2w = "2W" in avail

    # CAR: select report month row, but query range includes previous month
    rider_car  = pick_month_row(filter_all(q4659, city=city), start_date)
    driver_car = pick_month_row(filter_all(q6150, city=city), start_date)

    # promo (region-only): city + vtype
    q6030_2w = vt(q6030, "2W", city=city)
    q6030_3w = vt(q6030, "3W", city=city)
    q6030_4w = vt(q6030, "4W", city=city)

    general = {
        "unique_rider_searches": safe_val(q6145, "avg_daily_searches"),
        "book_search_ratio":     safe_val(q6145, "monthly_book_search_ratio"),
        "median_searched_fare":  safe_val(q6366, "median_searched_fare"),
        "non_trip_marketing":    None,
    }

    def build_vehicle(v):
        vlow = v.lower()

        # DEMAND: region-only -> MUST filter by city + vtype if city column exists
        # 6152 output has: month, city, vehicle_type
        dem = filter_all(q6152, city=city)          # keep only the city
        dem = vt(dem, v)                           # then pick vehicle_type
        dem = pick_month_row(dem, start_date)      # then pick report month (YYYY-MM-01)

        # SUPPLY/EFFICIENCY: filter city when possible
        ping = vt(q6151, v, city=city) if has_col(q6151, "city") else vt(q6151, v)
        onl  = vt(q6157, v, city=city) if has_col(q6157, "city") else vt(q6157, v)
        fin  = vt(q6149, v, city=city) if has_col(q6149, "city") else vt(q6149, v)
        uti  = vt(q6161, v, city=city) if has_col(q6161, "city") else vt(q6161, v)

        mr   = vt(q6147, v, city=city) if has_col(q6147, "city") else vt(q6147, v)
        ctr  = vt(q6144, v, city=city) if has_col(q6144, "city") else vt(q6144, v)
        ttm  = vt(q4509, v, city=city) if has_col(q4509, "city") else vt(q4509, v)
        eta  = vt(q6148, v, city=city) if has_col(q6148, "city") else vt(q6148, v)

        # fares are city-filtered by param already
        fare = vt(q6138, v)

        # fee is region-only -> filter by city (+ vtype if possible)
        if has_col(q6189, "vehicle_type") or has_col(q6189, "w_type"):
            fee = vt(q6189, v, city=city)
        else:
            fee = filter_rows(q6189, city=city)

        promo = {"2W": q6030_2w, "3W": q6030_3w, "4W": q6030_4w}.get(v, pd.DataFrame())

        return {
            "unique_bookings_daily":            safe_val(dem,  "unique_orders_daily_avg"),
            "completed_daily":                  safe_val(dem,  "completed_daily_avg"),
            "completion_rate":                  safe_val(dem,  "completion_rate"),

            "pinged_drivers_daily":             safe_val(ping, "avg_daily_drivers_gt1_ping"),
            "avg_online_hours":                 safe_val(onl,  "avg_daily_online_hours"),
            "completed_drivers":                safe_val(fin,  "monthly_unique_finished_drivers"),
            "ride_per_driver":                  safe_val(fin,  "ride_per_driver"),
            "driver_utilisation":               div(safe_val(uti, "avg_utilisation_hours"), safe_val(onl, "avg_daily_online_hours")),

            "match_rate":                       safe_val(mr,   "match_rate"),
            "first_try_cater_rate":             safe_val(ctr,  "first_try_cater_rate"),
            "median_time_to_match_sec":         safe_val(ttm,  "median_time_to_match_sec"),
            "median_eta":                       safe_val(eta,  "median_eta_minutes"),

            "median_booked_fare":               safe_val(fare, "median_booked_fare"),
            "median_completed_fare":            safe_val(fare, "median_completed_fare"),
            "median_matched_driver_earnings":   safe_val(fare, "median_matched_driver_earnings"),
            "median_completed_driver_earnings": safe_val(fare, "median_completed_driver_earnings"),

            "driver_incentive":                 None,
            "rider_promo_cost":                 safe_val(promo, "rider_promo_cost_per_trip"),
            "rider_promo_rate":                 safe_val(promo, "rider_promo_rate"),
            "total_spend":                      None,
            "system_fee":                       safe_val(fee,  "avg_system_fee_per_trip"),

            # CAR fields: be flexible on column names
            "rider_completed":                  first_present_val(rider_car,  [f"rider_completed_{vlow}", "rider_completed"]),
            "rider_activated":                  first_present_val(rider_car,  [f"rider_activated_{vlow}", "rider_activated"]),
            "rider_churn_rate":                 first_present_val(rider_car,  [f"rider_churn_rate_{vlow}", "rider_churn_rate"]),

            "driver_completed":                 first_present_val(driver_car, [f"driver_completed_{vlow}", "driver_completed"]),
            "driver_activated":                 first_present_val(driver_car, [f"driver_activated_{vlow}", "driver_activated"]),
            "driver_churn_rate":                first_present_val(driver_car, [f"driver_churn_rate_{vlow}", "driver_churn_rate"]),
        }

    d2w = build_vehicle("2W") if has_2w else {}
    d3w = build_vehicle("3W")
    d4w = build_vehicle("4W")

    return general, d2w, d3w, d4w, has_2w

# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    load_dotenv()

    redash = Redash(
        key=os.getenv("REDASH_API_KEY"),
        base_url=os.getenv("REDASH_BASE_URL")
    )

    date, start_date, end_date, churn_start, label = prev_month_info()

    print(f"KPI Tracker - {label}")
    print(f"  Period      : {start_date} -> {end_date}")
    print(f"  Churn range : {churn_start} -> {end_date}\n")

    wb = Workbook()
    temp_ws = wb.active
    temp_ws.title = "TEMP"

    tasks = [
        ("SG",        lambda: (fetch_sg(redash, date, start_date, end_date, churn_start),              "sg_hk")),
        ("HK",        lambda: (fetch_hk(redash, date, start_date, end_date, churn_start),              "sg_hk")),
        ("TH",        lambda: (fetch_th(redash, date, start_date, end_date, churn_start),              "th_vn")),
        ("VN-HCMC",   lambda: (fetch_vn(redash, start_date, end_date, churn_start, "HCM"),             "th_vn")),
        ("VN-Hanoi",  lambda: (fetch_vn(redash, start_date, end_date, churn_start, "HAN"),             "th_vn")),
        ("KH-PNH",    lambda: (fetch_kh_city(redash, start_date, end_date, churn_start, "PNH"),        "kh")),
        ("KH-OTHERS", lambda: (fetch_kh_city(redash, start_date, end_date, churn_start, "KH-OTHERS"),  "kh")),
    ]

    created = []
    for name, fn in tasks:
        print(f"-> {name}")
        try:
            result, layout = fn()
            ws = wb.create_sheet(name)

            if layout == "sg_hk":
                write_sheet(ws, label, rows_sg_hk(result))

            elif layout == "th_vn":
                gen, d2w, d4w = result
                write_sheet(ws, label, rows_th_vn(gen, d2w, d4w))

            else:
                gen, d2w, d3w, d4w, has_2w = result
                write_sheet(ws, label, rows_kh(gen, d2w, d3w, d4w, has_2w=has_2w))

            created.append(name)
            print("   done")

        except Exception as e:
            import traceback
            print(f"   FAILED: {e}")
            traceback.print_exc()

    if created:
        wb.remove(temp_ws)
    else:
        temp_ws.title = "ERROR"
        temp_ws["A1"] = "All regions failed - check logs above."

    output_file = f"KPI_Data_{label}.xlsx"
    wb.save(output_file)
    print(f"\nSaved: {output_file}")

    slack = SlackBot()
    slack.uploadFile(
        output_file,
        os.getenv("SLACK_CHANNEL"),
        f"KPI Data for {label.replace('_', ' ')}"
    )
    print("Uploaded to Slack")

    try:
        os.remove(output_file)
    except OSError:
        pass


if __name__ == "__main__":
    main()